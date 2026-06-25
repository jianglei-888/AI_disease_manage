"""医疗知识库 RAG。

提供：
1. MedicalRAG：单例，用于后台 /v1/backend/upload 入库 + /v1/backend/search 查询
2. retrieve_medical_context：纯函数，被 MedicalRAG 和 LangGraph 的 rag 节点共享
"""
import os
from typing import List

from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
    Docx2txtLoader,
    CSVLoader,
    UnstructuredPowerPointLoader,
)
from langchain_core.documents import Document
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_core.runnables import RunnableLambda, RunnablePassthrough
from langchain_openai import ChatOpenAI
from langchain_text_splitters import RecursiveCharacterTextSplitter
from pymilvus import MilvusClient

from app.utils.embeddings import DashScopeCompatibleEmbeddings
from config.settings import settings
from app.utils.logger import get_logger

logger = get_logger(__name__)


def _load_xlsx_with_openpyxl(file_path: str) -> List[Document]:
    """用 openpyxl 读 xlsx，绕开 unstructured 的传递依赖（networkx/msoffcrypto/...）。

    行为：每个 sheet 一行一条 Document；第一行作表头，prefix 到每条 page_content 前面，
    让"嘉瑟宜"这种药品名能跟它同行的"缬沙坦/分类/适应症"embed 到同一个 chunk 里。
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=True)
    docs: List[Document] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = ["" if c is None else str(c).strip() for c in rows[0]]
        for row in rows[1:]:
            cells = ["" if c is None else str(c) for c in row]
            # 跳过全空行
            if not any(c.strip() for c in cells):
                continue
            # 行内容：表头:值 表头:值 ...，比 "\t" 拼接更利于 embedding 抓取字段语义
            kv = [f"{h}：{v}" for h, v in zip(header, cells) if h and v.strip()]
            if not kv:
                # 没表头或全空值，回退到 tab 拼接
                kv = ["\t".join(cells)]
            page_content = f"[Sheet: {sheet_name}]\n" + "\n".join(kv)
            docs.append(Document(page_content=page_content, metadata={"source": file_path}))
    wb.close()
    return docs


def _load_xls_with_openpyxl(file_path: str) -> List[Document]:
    """xls 旧格式：openpyxl 不支持，仍走 unstructured（如果环境里有）。"""
    from langchain_community.document_loaders import UnstructuredExcelLoader

    return UnstructuredExcelLoader(file_path, mode="elements").load()


def retrieve_medical_context(query: str, top_k: int = 3) -> str:
    """共享的检索函数：query → Milvus 检索 → 拼接的 context 文本。

    返回纯文本（用 "\n\n" 拼接的若干片段），无相关结果时返回中文提示。
    """
    embeddings = DashScopeCompatibleEmbeddings()
    client = MilvusClient(uri=settings.MILVUS_URI, db_name=settings.MILVUS_DB)

    query_vector = embeddings.embed_query(query)
    res = client.search(
        collection_name=settings.MILVUS_COLLECTION,
        data=[query_vector],
        limit=top_k,
        output_fields=["text", "source"],
    )

    hits = res[0] if res else []
    if not hits:
        return "知识库中目前没有关于此问题的信息。"

    return "\n\n".join(hit["entity"]["text"] for hit in hits)


class MedicalRAG:
    """单例 RAG 封装。后台模板页用。"""

    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def _ensure_init(self):
        if self._initialized:
            return
        self._initialized = True
        self.embeddings = DashScopeCompatibleEmbeddings()
        self.llm = ChatOpenAI(
            model=settings.LLM_MODEL,
            api_key=settings.LLM_API_KEY,
            base_url=settings.LLM_BASE_URL,
        )
        self.vector_db = MilvusClient(
            uri=settings.MILVUS_URI,
            db_name=settings.MILVUS_DB,
            collection_name=settings.MILVUS_COLLECTION,
        )
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.CHUNK_SIZE,
            chunk_overlap=settings.CHUNK_OVERLAP,
            separators=["\n\n", "\n", "。", "，", " "],
        )
        self.prompt = PromptTemplate(
            template=(
                "你是专业医疗助手，必须严格依据知识库回答，禁止编造。\n"
                "若无相关信息，回复：暂无相关医疗知识，请咨询专业医生。\n\n"
                "知识库：\n{context}\n\n"
                "问题：{question}\n"
                "回答："
            ),
            input_variables=["context", "question"],
        )

    def load_file(self, file_path: str) -> int:
        """加载一个文档入库，返回切分后的块数。"""
        self._ensure_init()

        ext = os.path.splitext(file_path)[-1].lower()
        loader_map = {
            ".pdf": lambda p: PyPDFLoader(p).load(),
            ".docx": lambda p: Docx2txtLoader(p).load(),
            ".xlsx": lambda p: _load_xlsx_with_openpyxl(p),
            ".xls": lambda p: _load_xls_with_openpyxl(p),
            ".csv": lambda p: CSVLoader(p, encoding="utf-8").load(),
            ".txt": lambda p: TextLoader(p, encoding="utf-8").load(),
            ".md": lambda p: TextLoader(p, encoding="utf-8").load(),
            ".pptx": lambda p: UnstructuredPowerPointLoader(p).load(),
        }
        if ext not in loader_map:
            raise ValueError(f"不支持的文件格式：{ext}")

        docs = loader_map[ext](file_path)
        # xlsx 已经是一行一条 Document（短小），text_splitter 不会再切；
        # 其它格式仍按 chunk_size 切
        if ext == ".xlsx":
            splits = docs
        else:
            splits = self.text_splitter.split_documents(docs)

        # 同一文件重复上传时先清掉旧 record，避免嘉瑟宜被灌 3 份这种问题。
        # 用 basename 归一化：库里的 source 统一是文件名，跟前端上传时给的路径无关。
        # 历史库里有用绝对路径/正反斜杠写过的脏数据，也一并清掉。
        source = os.path.basename(file_path)
        try:
            existing = self.vector_db.query(
                collection_name=settings.MILVUS_COLLECTION,
                filter=f'source == "{source}"',
                output_fields=["id"],
                limit=10000,
            )
            if existing:
                ids = [r["id"] for r in existing]
                self.vector_db.delete(
                    collection_name=settings.MILVUS_COLLECTION,
                    ids=ids,
                )
                logger.info("按 basename 命中旧记录：%s, %d 条", source, len(ids))
        except Exception as e:
            logger.warning("清旧记录失败（不影响入库）：%s", e)

        # 顺手清一下历史脏数据：v0.2 之前 source 没归一化，用过绝对路径或正反斜杠相对路径
        # 的旧 record。用 filter 按字符串筛反斜杠不靠谱，索性拉全表本地过滤 basename。
        try:
            all_recs = self.vector_db.query(
                collection_name=settings.MILVUS_COLLECTION,
                output_fields=["id", "source"],
                limit=16384,
            )
            dirty_ids = [r["id"] for r in all_recs
                         if r["source"] != source
                         and os.path.basename(r["source"]) == source]
            if dirty_ids:
                self.vector_db.delete(
                    collection_name=settings.MILVUS_COLLECTION,
                    ids=dirty_ids,
                )
                logger.info("清掉历史脏 source：%s, %d 条", source, len(dirty_ids))
        except Exception as e:
            logger.warning("清历史脏 source 失败（不影响入库）：%s", e)

        data = [
            {
                "vector": self.embeddings.embed_query(doc.page_content),
                "text": doc.page_content,
                "source": source,
            }
            for doc in splits
        ]
        self.vector_db.insert(collection_name=settings.MILVUS_COLLECTION, data=data)
        logger.info("入库完成：%s, 共 %d 块", file_path, len(splits))
        return len(splits)

    def get_chain(self):
        """构建 RAG 问答 chain（后台查询用）。"""
        self._ensure_init()
        return (
            {
                "context": RunnableLambda(retrieve_medical_context),
                "question": RunnablePassthrough(),
            }
            | self.prompt
            | self.llm
            | StrOutputParser()
        )


# 单例：被 backend/views.py 直接调用
rag = MedicalRAG()